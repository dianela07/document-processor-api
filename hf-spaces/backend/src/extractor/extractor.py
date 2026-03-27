"""
Módulo de extracción de datos de múltiples formatos de archivo.
Soporta: PDF, CSV, Excel, TXT/Logs
"""
import io
import csv
import re
from typing import Union
from pathlib import Path
import logging

logger = logging.getLogger(__name__)


class FileExtractor:
    """Extractor universal de contenido de archivos."""
    
    SUPPORTED_FORMATS = {
        'pdf': ['.pdf'],
        'csv': ['.csv'],
        'excel': ['.xlsx', '.xls'],
        'text': ['.txt', '.log']
    }
    
    def __init__(self):
        self._extractors = {
            'pdf': self._extract_pdf,
            'csv': self._extract_csv,
            'excel': self._extract_excel,
            'text': self._extract_text
        }
    
    def extract(self, file_bytes: bytes, filename: str) -> dict:
        """
        Extrae contenido de un archivo según su extensión.
        
        Args:
            file_bytes: Contenido del archivo en bytes
            filename: Nombre del archivo con extensión
            
        Returns:
            dict con 'content', 'format', 'metadata'
        """
        extension = Path(filename).suffix.lower()
        file_type = self._get_file_type(extension)
        
        if not file_type:
            raise ValueError(f"Formato no soportado: {extension}")
        
        logger.info(f"Extrayendo contenido de {filename} (tipo: {file_type})")
        
        extractor = self._extractors[file_type]
        content, metadata = extractor(file_bytes)
        
        return {
            'content': content,
            'format': file_type,
            'extension': extension,
            'filename': filename,
            'metadata': metadata
        }
    
    def _get_file_type(self, extension: str) -> Union[str, None]:
        """Determina el tipo de archivo por extensión."""
        for file_type, extensions in self.SUPPORTED_FORMATS.items():
            if extension in extensions:
                return file_type
        return None
    
    def _extract_pdf(self, file_bytes: bytes) -> tuple[str, dict]:
        """Extrae texto de un PDF."""
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(file_bytes))
            texts = []
            for page in reader.pages:
                texts.append(page.extract_text() or "")
            
            metadata = {
                'pages': len(reader.pages),
                'info': dict(reader.metadata) if reader.metadata else {}
            }
            return "\n".join(texts), metadata
        except Exception as e:
            logger.error(f"Error extrayendo PDF: {e}")
            raise
    
    def _extract_csv(self, file_bytes: bytes) -> tuple[str, dict]:
        """Extrae contenido de un CSV como texto estructurado."""
        try:
            content = file_bytes.decode('utf-8')
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
            
            # Convertir a texto legible para LLM
            text_lines = []
            for i, row in enumerate(rows):
                text_lines.append(f"Fila {i+1}: " + ", ".join(f"{k}={v}" for k, v in row.items()))
            
            metadata = {
                'rows': len(rows),
                'columns': reader.fieldnames or []
            }
            return "\n".join(text_lines), metadata
        except Exception as e:
            logger.error(f"Error extrayendo CSV: {e}")
            raise
    
    def _extract_excel(self, file_bytes: bytes) -> tuple[str, dict]:
        """Extrae contenido de un archivo Excel."""
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            
            all_text = []
            sheets_info = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows_data = []
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = [str(cell) if cell is not None else "" for cell in row]
                    if any(row_text):  # Solo filas con contenido
                        rows_data.append(", ".join(row_text))
                
                all_text.append(f"=== Hoja: {sheet_name} ===")
                all_text.extend(rows_data)
                sheets_info[sheet_name] = len(rows_data)
            
            metadata = {
                'sheets': workbook.sheetnames,
                'rows_per_sheet': sheets_info
            }
            return "\n".join(all_text), metadata
        except Exception as e:
            logger.error(f"Error extrayendo Excel: {e}")
            raise
    
    def _extract_text(self, file_bytes: bytes) -> tuple[str, dict]:
        """Extrae contenido de archivos de texto/logs."""
        try:
            content = file_bytes.decode('utf-8')
            lines = content.split('\n')
            
            # Detectar patrones comunes en logs
            error_count = len(re.findall(r'\b(ERROR|FATAL|CRITICAL)\b', content, re.IGNORECASE))
            warning_count = len(re.findall(r'\b(WARNING|WARN)\b', content, re.IGNORECASE))
            
            metadata = {
                'lines': len(lines),
                'characters': len(content),
                'errors_detected': error_count,
                'warnings_detected': warning_count
            }
            return content, metadata
        except Exception as e:
            logger.error(f"Error extrayendo texto: {e}")
            raise
    
    @classmethod
    def get_supported_extensions(cls) -> list[str]:
        """Retorna lista de todas las extensiones soportadas."""
        extensions = []
        for ext_list in cls.SUPPORTED_FORMATS.values():
            extensions.extend(ext_list)
        return extensions
