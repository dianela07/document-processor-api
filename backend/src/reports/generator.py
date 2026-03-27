"""
Generador de reportes automáticos.
Demuestra: Automatización, Python scripting, generación de documentos.
"""
import io
import csv
import json
from datetime import datetime
from typing import Any
from pathlib import Path
import logging

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Genera reportes en múltiples formatos."""
    
    def __init__(self):
        self.output_dir = Path("reports")
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate(
        self,
        data: list[dict],
        format: str = "excel",
        title: str = "Reporte",
        include_metadata: bool = True
    ) -> bytes:
        """
        Genera un reporte en el formato especificado.
        
        Args:
            data: Lista de diccionarios con los datos
            format: Formato de salida (excel, csv, json)
            title: Título del reporte
            include_metadata: Si incluir metadatos del reporte
            
        Returns:
            Contenido del reporte en bytes
        """
        logger.info(f"Generando reporte '{title}' en formato {format}")
        
        generators = {
            'excel': self._generate_excel,
            'csv': self._generate_csv,
            'json': self._generate_json
        }
        
        if format not in generators:
            raise ValueError(f"Formato no soportado: {format}")
        
        return generators[format](data, title, include_metadata)
    
    def _generate_excel(
        self,
        data: list[dict],
        title: str,
        include_metadata: bool
    ) -> bytes:
        """Genera reporte en formato Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl no instalado, generando CSV en su lugar")
            return self._generate_csv(data, title, include_metadata)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título del reporte
        ws.merge_cells('A1:E1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Metadatos
        if include_metadata:
            ws['A2'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A3'] = f"Total registros: {len(data)}"
            start_row = 5
        else:
            start_row = 3
        
        if not data:
            ws[f'A{start_row}'] = "No hay datos para mostrar"
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
        
        # Headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Datos
        for row_idx, row_data in enumerate(data, start_row + 1):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                # Convertir diccionarios a JSON string
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # Ajustar ancho de columnas
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=start_row + 1, max_col=col_idx, min_col=col_idx):
                for cell in row:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)[:50]))
                    except:
                        pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        logger.info(f"Reporte Excel generado: {len(data)} registros")
        return output.getvalue()
    
    def _generate_csv(
        self,
        data: list[dict],
        title: str,
        include_metadata: bool
    ) -> bytes:
        """Genera reporte en formato CSV."""
        output = io.StringIO()
        
        if include_metadata:
            output.write(f"# {title}\n")
            output.write(f"# Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            output.write(f"# Total registros: {len(data)}\n")
            output.write("\n")
        
        if not data:
            output.write("No hay datos\n")
            return output.getvalue().encode('utf-8')
        
        headers = list(data[0].keys())
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        
        for row in data:
            # Convertir valores complejos a string
            clean_row = {}
            for k, v in row.items():
                if isinstance(v, (dict, list)):
                    clean_row[k] = json.dumps(v, ensure_ascii=False)
                else:
                    clean_row[k] = v
            writer.writerow(clean_row)
        
        logger.info(f"Reporte CSV generado: {len(data)} registros")
        return output.getvalue().encode('utf-8')
    
    def _generate_json(
        self,
        data: list[dict],
        title: str,
        include_metadata: bool
    ) -> bytes:
        """Genera reporte en formato JSON."""
        report = {
            "data": data
        }
        
        if include_metadata:
            report["metadata"] = {
                "title": title,
                "generated_at": datetime.now().isoformat(),
                "total_records": len(data)
            }
        
        logger.info(f"Reporte JSON generado: {len(data)} registros")
        return json.dumps(report, indent=2, ensure_ascii=False, default=str).encode('utf-8')
    
    def save_report(
        self,
        content: bytes,
        filename: str,
        format: str
    ) -> Path:
        """Guarda el reporte en disco."""
        extension_map = {'excel': '.xlsx', 'csv': '.csv', 'json': '.json'}
        ext = extension_map.get(format, '.txt')
        
        filepath = self.output_dir / f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
        
        with open(filepath, 'wb') as f:
            f.write(content)
        
        logger.info(f"Reporte guardado: {filepath}")
        return filepath
